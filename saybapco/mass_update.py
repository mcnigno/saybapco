import glob, os, shutil, uuid
from app import models, db
from flask_appbuilder import filemanager
import uuid, openpyxl
from helpers import sanetext
from config import UPLOAD_FOLDER
from flask_appbuilder.filemanager import get_file_original_name
import openpyxl
from flask import send_file
from config import UPLOAD_FOLDER
 
errors_list = set()
basedir = os.path.abspath(os.path.dirname(__file__))


def test_closed(doc_id):
    doc = models.Document
    document = db.session.query(doc).filter(doc.id==doc_id).first()
    print('Check Doc Function')
    
    closed = True
    for com in document.comments:
        
        if com.closed == False:
            closed = False
            break
        
        print('comment closed:',com, com.closed)
    print('comment closed:',com, com.closed)
    document.closed = closed
    print('document closed?', document.closed)
    document.changed_by_fk = '1'
    db.session.commit()
def report_all():
    filepath = UPLOAD_FOLDER + 'report/cs_dashboard.xlsx'
    workbook = openpyxl.load_workbook(filepath)
    
    comment = models.Comments
    revision = models.Revisions
    document = models.Document

    wc = workbook['Comments']    
    wr = workbook['Revisions']
    wd = workbook['Documents']

    # 
    # Populate Comments Sheet
    #

    '''
    comments = db.session.query(comment).all()
    
    row = 1
    col = 1 

    for comment in comments:
        print('comments *********************')
        _ = wc.cell(column=1, row=row, value=str(comment.id))
        _ = wc.cell(column=2, row=row, value=str(comment.partner))
        _ = wc.cell(column=3, row=row, value=str(comment.document))
        _ = wc.cell(column=4, row=row, value=str(comment.revision))
        _ = wc.cell(column=5, row=row, value=str(comment.included))
        _ = wc.cell(column=6, row=row, value=str(comment.closed))

        row += 1
    '''
    # 
    # Populate revisions Sheet
    #

    
    revisions = db.session.query(revision).all()
    
    row = 1
    col = 1 

    for revision in revisions:
        _ = wr.cell(column=1, row=row, value=str(revision.id))
        _ = wr.cell(column=2, row=row, value=str(revision.file))
        _ = wr.cell(column=3, row=row, value=str(revision.revision))
        _ = wr.cell(column=4, row=row, value=str(revision.trasmittal))
        _ = wr.cell(column=5, row=row, value=str(revision.date_trs))
        _ = wr.cell(column=6, row=row, value=str(revision.document))
        _ = wr.cell(column=7, row=row, value=str(revision.reply))

        row += 1

    # 
    # Populate documents Sheet
    #

    
    documents = db.session.query(document).all()
    
    row = 2
    col = 1 
    _ = wd.cell(column=1, row=1, value='ID')
    _ = wd.cell(column=2, row=1, value='Bapco Code')
    _ = wd.cell(column=3, row=1, value='Partner')
    _ = wd.cell(column=4, row=1, value='Revisions')
    _ = wd.cell(column=5, row=1, value='Open CS')

    for document in documents:
        _ = wd.cell(column=1, row=row, value=str(document.id))
        _ = wd.cell(column=2, row=row, value=str(document.name()))
        _ = wd.cell(column=3, row=row, value=str(document.partner))
        _ = wd.cell(column=4, row=row, value=str(document.revision))
        _ = wd.cell(column=5, row=row, value=str(document.count_open()))
        
        

        row += 1
        


    #for col in wc.iter_cols(min_row=1, max_col=3, max_row=3):
        #for cell in col:
            #cell.value = 'hello world'
    workbook.save(filepath) 

    
    
    
    file = UPLOAD_FOLDER + 'report/cs_dashboard.xlsx'
    #print(file.name)
    return file
    


def transmittall():
    print('transmittal function ********')

    #
    # Update Trasmittal and date matching data from this two report
    #   
    report = open(basedir+'/report/Report.xlsx', mode='rb')
    wb = openpyxl.load_workbook(report)
    ws = wb.active
 
    report_bapco = open(basedir+'/report/Report_by_BAPCO.xlsx', mode='rb')
    bb = openpyxl.load_workbook(report_bapco)
    bs = bb.active
    
    #
    # For every document's revision in the db look for data
    #

    document = models.Document
    document_l = db.session.query(document).all()
    
    for doc in document_l:
        code = '-'.join([doc.unit, doc.materialclass, doc.doctype, doc.serial, doc.sheet])
        for rev in doc.revision:

            #print(code, '*', rev)
            # 
            # Over Report.xlsx Column[2] == code AND Column[4] == revision
            #
            for row in ws.iter_rows(min_row=2):
                #print('------------')
                #print(row[2].value, code)
                #print(row[4].value, rev)
                if row[2].value == code and row[4].value == str(rev):
                      

                    bapco_code = row[2].value
                    revision = row[4].value
                    bapco_transmittal = row[0].value
                     
                    for nrow in bs.iter_rows(min_row=2):
                        if bapco_transmittal == nrow[3].value[6:]:
                            
                            print(code)
                            print(nrow[3].value[6:])
                            print(nrow[0].value)
                            trans = nrow[0].value
                            date = str(nrow[1].value)
                            gg,mm,aa = date.split("/")
        
                            date = aa + r"/" + mm + r"/" + gg 
        
                            revisions = models.Revisions
                            #rev_update = db.session.query(revisions).filter(revisions.id == rev.id).first()
                            rev.trasmittal = trans
                            rev.changed_by_fk = '1'
                            rev.date_trs = date
                            rev.note = nrow[3].value
                            doc.changed_by_fk = '1'
                            doc.partner = trans[4:7]

                            db.session.commit() 
    db.session.close()
                         


                           
def reply_rev():
    '''
    Danilo,
    partendo dalla lista Report.xlsx:
    '''
    report = open(UPLOAD_FOLDER + 'report/Report.xlsx', mode='rb')
    wb = openpyxl.load_workbook(report)
    ws = wb.active
    '''
    quando la revisione del documento1 è X-REPLY (X=numero/lettera) cerca nella lista documento1 con rev. X
    '''
    document = models.Document
    document_l = db.session.query(document).filter()

    for doc in document_l:
        code = '-'.join([doc.unit, doc.materialclass, doc.doctype, doc.serial, doc.sheet])
        for rev in doc.revision:
            if rev.reply:
                print('is a a reply:', doc, rev)
    '''
    Se nella linea precedente il numero di documento corrisponde (documento1 con rev ≠ X) , prendi il transmittal e la data corrispondenti.
    Spero di essere stata chiara.
    Federica
    '''





def comments(item): 

    print('file processed:', str(item.file))
    file = open(basedir + '/batch/' + get_file_original_name(item.file), mode='rb')
    
    #print('file processed:', file)

    #file = open('/Users/dp/py3/saybapco/saybapco/comments/019-C-GAD-10010-001-RAP-CS REPLY.xlsx', mode='rb')
    #filename = '019-C-GAD-10010-001-RAP-CS REPLY.xlsx'

    filename = str(item.file)
    unit = filename[0:3]

    #partner = filename[4]

    mat = filename[4]
    doctype = filename[6:9]
    serial = filename[10:15]
    type_reply = False
    print('Reply identification: ', filename[-8:-5] )
    try:
        #if filename[26:29] == "REP":
        if filename[-8:-5] == "REP":
            print('Reply identification: ', filename[-8:-5] )
            type_reply = True
    except:
        pass
    
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    session = db.session
    comments= models.Comments
    documents = models.Document
    revisions = models.Revisions

    revision = session.query(revisions).filter(revisions.revision == item.revision, revisions.document_id == item.document_id).first()
    #revision = session.query(revisions).filter(revisions.id == item.id).first()
    session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).delete()
    
    try:

        print('doc id',item.document_id,'rev id', revision.id, 'item id', item.id)
        
        comm_list = session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).all()
        print('the comments len is:', len(comm_list))
        
        session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).delete()
        print('delete query executed')
    
    except:

        print('some problem here')
        pass
    

    #session.commit()
    #check columns label
    #for row in ws.iter_colum()
    #print('document id', item.document_id,'revision:', item.id)
    partner = item.partner
    #revision = item.revision
    
    for row in ws.iter_rows(min_row=2):
        
        if row[0].value is not None:
            #print(row[0].value, row[1].value, row[2].value,
              #row[3].value, row[4].value, row[5].value, row[6].value, row[7].value)
            try:


                id_c = sanetext(row[0].value)
                style = sanetext(row[1].value)
                page = sanetext(row[2].value)
                author = sanetext(row[3].value)
                comment = sanetext(row[4].value)
                reply = sanetext(row[5].value)
                included = sanetext(row[6].value)
                closed = sanetext(row[7].value)


                print(id_c, style, author)
                
                yes = ['Y','y','Yes','YES']
                if closed in yes:
                    closed = True
                else:
                    closed = False
                
                if included in yes:
                    included = True
                else:
                    included = False

                print('before comment')
                print('document id', item.document_id,'revision:', item.id)
                comm = comments(created_by_fk = 1, changed_by_fk = 1, id_c=id_c, partner=partner, style=style, page=page, author=author, comment=comment,
                                reply=reply, included=included, closed=closed,
                                document_id=item.document_id, revision_id=item.id, type_reply=type_reply)
                
                session.add(comm)

                print('after comment')
            
                
            except:
                print('something is None')
                errors_list.add(str(item.file))
                pass
    
    session.commit()
    print('list of Errors:')
    print(errors_list)
    return 'done'

def check_doc_closed(doc_id):
    doc = models.Document
    document = db.session.query(doc).filter(doc.id==doc_id).first()
    print('Check Doc Function')
    closed = True
    try:
        for com in document.comments:
            print(com, com.closed)
            if com.closed == False:
                closed = False
                break
            print('comment closed:',com, com.closed)
    except:
        closed = False
    document.closed = closed
    document.changed_by_fk = '1'
    db.session.commit()

def check_Doc(item):
    filename = item
    unit = filename[0:3]
    partner = "ND"
    mat = filename[4]
    doctype = filename[6:9]
    serial = filename[10:15]
    sheet = filename[16:19]

    
    session = db.session
    doc = models.Document
    document = session.query(doc).filter(doc.unit == unit, doc.materialclass == mat,
                                 doc.doctype == doctype, doc.serial == serial, doc.sheet == sheet).first()

    if document:
        
        return document.id, document.partner
    else:
        document = doc(created_by_fk = 1, changed_by_fk = 1, unit=unit, materialclass=mat, doctype=doctype, 
                       serial=serial, sheet=sheet, partner=partner)
        
        session.add(document)

        session.flush()
        session.commit()

        return document.id, document.partner


def mass_update():
    
    os.chdir(basedir +'/batch')

    rev_order = ['A','B','C','D','E','F','G','H','I','L','M','N','O','P','Q','R','S','T',
                'U','V','Z','0','1','2','3','4','5','6','7','8','9','10']
    
    for revision in rev_order:
        #print('processing Rev:', revision)
        # Check the revision A ...
        i = 0
        for file in glob.glob("*.xlsx"):
            i += 1
            print(i, file)

            
            _, _, _, _, _, rev, cs = file.split('-')
            cs = cs[:-5]
            #print('rev:', rev,'cs:', cs)
            
            if rev[1] == revision and cs != 'CS REP':
                file_sec = filemanager.secure_filename(file)
                #print(file_sec)
                file_byte = open(file)
                file_uuid = str(uuid.uuid4()) + '_sep_' + file_byte.name

                if os.path.exists(file):
                    src = os.path.realpath(file)
                
                    #head, tail = os.path.split(src)

                    dst = UPLOAD_FOLDER + file_uuid

                    shutil.copy(src, dst)
            


                #print('FILE BYTE', file_byte) 
                #filemanager.FileManager.save_file(file_byte, file_sec)
                new_rev = models.Revisions(created_by_fk = 1, changed_by_fk = 1, file=file_uuid, revision = rev[1:], trasmittal = 'to update', date_trs = '2015-01-01')
                new_rev.document_id, new_rev.partner = check_Doc(file)
                
                #new_rev.revision = revision
                rev_to_check = models.Revisions
                exist = db.session.query(rev_to_check).filter(rev_to_check.document_id == new_rev.document_id, rev_to_check.revision == rev[1:]).first()
                if exist is not None:
                    print('This Rev Already Exist - Not Added')
                else:
                    print('exist is there..', exist)
                    db.session.add(new_rev)


                    db.session.flush()
                    comments(new_rev)
                    check_doc_closed(new_rev.document_id)
        
        db.session.commit()
        db.session.close()
        # Check for the replay
                
        for file in glob.glob("*.xlsx"):
            _, _, _, _, _, rev, cs = file.split('-')
            cs = cs[:-5]
            if rev[1] == revision and cs == 'CS REP':
                file_sec = filemanager.secure_filename(file)
                #print('******************* REPLY')
                #print(file_sec)
                file_byte = open(file, mode='rb') 
                file_uuid = str(uuid.uuid4()) + '_sep_' + file_byte.name
                #filemanager.FileManager.save_file(file_byte, file_sec)
                new_rev = models.Revisions(created_by_fk = 1, changed_by_fk = 1, file=file_uuid, revision = rev[1:], trasmittal = 'to update', date_trs = '2015-01-01')
                new_rev.document_id, new_rev.partner = check_Doc(file) 
                #print('the new rev id is:', new_rev.document_id)
                
                #new_rev.revision = revision
                new_rev.reply = True
                exist = db.session.query(models.Revisions).filter(rev_to_check.document_id == new_rev.document_id, rev_to_check.revision == rev[1:], rev_to_check.reply == True).first()
                if exist is not None:
                    print('This Rev Already Exist - Not Added')
                else:
                    print('exist is there..', exist)
                    db.session.add(new_rev)
                    db.session.flush()
                    comments(new_rev)
                    check_doc_closed(new_rev.document_id)

                #print(file)
            
                #print('REV NOT FOUND', rev[1], file)
        db.session.commit()
        db.session.close()
    


