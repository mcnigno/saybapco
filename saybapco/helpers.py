import openpyxl, re
from config import UPLOAD_FOLDER
from app import db, models
from flask_appbuilder.filemanager import get_file_original_name
from openpyxl.styles import Font, Color
from openpyxl.worksheet.table import Table, TableStyleInfo
import uuid
#from views import Report
from flask import render_template, abort
from urllib.parse import urlparse, parse_qs
import json, requests


rev_order = ['A','B','C','D','E','F','G','H','I','L','M','N','O','P','Q','R','S','T',
                'U','V','Z','0','1','2','3','4','5','6','7','8','9','10']

def set_position():
    documents_list = db.session.query(models.Document).all()[:2] 
    last_revision_list = []
    revision = models.Revisions
    comments = models.Comments
    session = db.session
    
    for doc in documents_list:
        revisions_list = db.session.query(revision).filter(revision.document_id == doc.id).order_by((revision.date_trs).asc()).all()
        count = 0
        #revision.p
        #
        # Set the revision position by transmittal date
        #
        for rev in revisions_list:
            rev.changed_by_fk = '1'
            rev.current = False
            rev.pos = count
            count += 1
    
        #session.commit()
        #db.session.close() 

    
        #
        # Set the last revision as current
        # 
        #   
        try:
            last_rev = revisions_list[-1]
        except:
            print('last rev index problem on', doc)
        
        
        last_rev.current = True
        last_rev.changed_by_fk = '1'
        
        # 
        # Commit to db
        #  

        
        #last_revision_list.append(last_rev)
        db.session.query(comments).filter(comments.document_id == last_rev.document_id).delete()
        
        db.session.commit() 
        last_revision_list.append(last_rev)
    
    set_last_rev_comments(last_revision_list)
        
def set_last_rev_comments(last_revision_list):
    session = db.session
    text = models.Comments

    for item in last_revision_list: 
        filename = get_file_original_name(item.file)
        
        type_reply = False

        if filename[-8:-5] == "REP":
            print('Reply identification: ', filename[-8:-5] )
            type_reply = True
        
        try:
            print(UPLOAD_FOLDER)
            file = open(UPLOAD_FOLDER + item.file, mode='rb')

            print(UPLOAD_FOLDER)
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            
            comment= models.Comments
            documents = models.Document
            revisions = models.Revisions

            #revision = session.query(revisions).filter(revisions.revision == item.revision, revisions.document_id == item.document_id).first()
            #revision = session.query(revisions).filter(revisions.id == item.id).first()
            #session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).delete()
        
            #print('doc id',item.document_id,'rev id', revision.id, 'item id', item.id)
            #comm_list = session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).all()
            #print('the comments len is:', len(comm_list))
            #session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).delete()
            #print('delete query executed')
            
        except:
            abort(400,'OPEN FILE - Please check your Excel file format.')
        

        #session.commit()
        #check columns label
        #for row in ws.iter_colum()
        #print('document id', item.document_id,'revision:', item.id)
        
        partner = "XXX"
        
        print('CS File', filename)
        #revision = item.revision
        try:
            
            for row in ws.iter_rows(min_row=2):
                #print(row[0].value, row[1].value, row[2].value,
                    #row[3].value, row[4].value, row[5].value, row[6].value, row[7].value)
                
                if row[0].value is not None:
                    #print('row 0 in not null', row[0].value)
                    


                    id_c = sanetext(row[0].value)
                    style = sanetext(row[1].value)
                    page = sanetext(row[2].value)
                    author = sanetext(row[3].value)
                    comment = sanetext(row[4].value)
                    reply = sanetext(row[5].value)
                    included = sanetext(row[6].value)
                    closed = sanetext(row[7].value)


                    #print(id_c, style, author)
                    


                    if closed == 'Y':
                        closed = True
                    else:
                        closed = False
                    
                    if included == 'Y':
                        included = True
                        closed = True
                    else:
                        included = False
                                    
                    #print('before comment')
                    print('document id', item.document_id,'revision:', item.id)
                    print(item.document_id, item.id, type_reply, comment)
                    #reply = True
                    
                    comm = text(id_c=id_c, partner=partner, style=style, page=page, author=author, comment=comment,
                                    reply=reply, included=included, closed=closed,
                                    document_id=item.document_id, revision_id=item.id, type_reply=type_reply)
                    comm.changed_by_fk = '1'
                    comm.created_by_fk = '1'
                    #print('before add')
                    print('reply: ', reply, 'partner', partner)
                    session.add(comm)
            
                    #session.commit()
                    #print('after comment')
                

                    
            
            print('CS COMMIT')
                
                    
        except:
            print('except for some reasons')
            abort(400,'AFTER COMMIT - Please check your Excel file format.')
        
        
        session.commit()
        print('CS Commit DONE')
     

def parse_escaped_character_match(match):
    return chr(int(match.group(1), 16))

def sanetext(input_string):
    if input_string is not None:
        return re.sub("_x([0-9A-F]{4})_", parse_escaped_character_match, str(input_string))
    return " "


#
#  Add Revisions uploading a Comment file by Bluebin.
#


def check_duplicates():
    revision = models.Revisions
    rev_list = db.session.query(revision).all()
    
    seen = {}
    dupes = []

    for rev in rev_list:
        rev_key = str(rev.document) + str(rev.revision)

        if rev.reply:
            rev_key = str(rev.document) + str(rev.revision) + "-Reply"

            
        
        if rev_key not in seen:
            seen[rev_key] = 1
        else:
            if seen[rev_key] == 1:
                dupes.append(rev_key)
            seen[rev_key] += 1
    print('seen',len(seen))
    print('dupes',len(dupes))
    for s in range(10):
        for s in seen:
            print(s)
    print('dupes', dupes)

def comments(item):
    revision = models.Revisions
    doc = item.document
    revisions_list = db.session.query(revision).filter(revision.document_id == doc.id).order_by((revision.date_trs).asc()).all()
    count = 0
    #revision.p
    #
    # Set the revision position by transmittal date
    #
    for rev in revisions_list:
        #rev.changed_by_fk = '1'
        rev.current = False
        rev.pos = count
        count += 1
    # 
    # Set the last revision by trans. date as current
    #

    this_last_rev = revisions_list[-1]
    this_last_rev.current = True

    comments = models.Comments
    #db.session.query(comments).filter(comments.document_id == item.document_id).delete()

    
    # old comment function
    
    try:
        print('file processed:', str(item.file))
        
        #print('file processed:', file)

        #file = open('/Users/dp/py3/saybapco/saybapco/comments/019-C-GAD-10010-001-RAP-CS REPLY.xlsx', mode='rb')
        #filename = '019-C-GAD-10010-001-RAP-CS REPLY.xlsx'

        filename = get_file_original_name(item.file)
        unit = filename[0:3]

        #partner = filename[4]

        mat = filename[4]
        doctype = filename[6:9]
        serial = filename[10:15]
        type_reply = False

        
        #if filename[26:29] == "REP":
        if filename[-8:-5] == "REP":
            print('Reply identification: ', filename[-8:-5] )
            type_reply = True
            
            
    except:
        abort(400,'Please check your file name.')
    
    try:
        print(UPLOAD_FOLDER)
        file = open(UPLOAD_FOLDER + item.file, mode='rb')

        print(UPLOAD_FOLDER)
        wb = openpyxl.load_workbook(file)
        ws = wb.active

        session = db.session
        comments= models.Comments
        #documents = models.Document
        #revisions = models.Revisions

        #revision = session.query(revisions).filter(revisions.revision == item.revision, revisions.document_id == item.document_id).first()
        #revision = session.query(revisions).filter(revisions.id == item.id).first()
        #session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).delete()
    
        #print('doc id',item.document_id,'rev id', revision.id, 'item id', item.id)
        #comm_list = session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).all()
        #print('the comments len is:', len(comm_list))
        #session.query(comments).filter(comments.document_id == item.document_id, comments.revision_id == revision.id).delete()
        #print('delete query executed')
        
    except:
        abort(400,'Please check your Excel file format.')
    

    #session.commit()
    #check columns label
    #for row in ws.iter_colum()
    #print('document id', item.document_id,'revision:', item.id)
    print('before the partner')
    partner = 'ND'
    print('here')
    #revision = item.revision
    try:
        
        for row in ws.iter_rows(min_row=2):
            #print(row[0].value, row[1].value, row[2].value,
                #row[3].value, row[4].value, row[5].value, row[6].value, row[7].value)
            
            if row[0].value is not None:
                #print('row 0 in not null', row[0].value)
                


                id_c = sanetext(row[0].value)
                style = sanetext(row[1].value)
                page = sanetext(row[2].value)
                author = sanetext(row[3].value)
                comment = sanetext(row[4].value)
                reply = sanetext(row[5].value)
                included = sanetext(row[6].value)
                closed = sanetext(row[7].value)


                #print(id_c, style, author)
                


                if closed == 'Y':
                    closed = True
                else:
                    closed = False
                
                if included == 'Y':
                    included = True
                    closed = True
                else:
                    included = False
                                
                print('before comment')
                print('document id', item.document_id,'revision:', item.id, partner)
                comm = comments(id_c=id_c, partner=partner, style=style, page=page, author=author, comment=comment,
                                reply=reply, included=included, closed=closed,
                                document_id=item.document_id, revision_id=item.id, type_reply=type_reply)
                
                session.add(comm)

                print('after comment')
            
                
    except:
        print()
        abort(400,'Please check your Excel file format.')
    
    
    session.commit()
    print('comments done')
    
    
    #session.close()
    
    return 'done'

def last_rev(self, item):

    # trova tutte le revisioni per il documento
    doc = item.document
    rev = item.revision
    print('last rev function:')
    print('document ID', doc.id)
    print('revisions', doc.revision)
    print('doc rev', rev)
    print('is a reply?', item.reply)
    #revisions = db.session.query(models.Revisions).filter(models.Revisions.document_id == item.id).all()
    #print('revision list', revisions)
    cur_rev_index = rev_order.index(rev)
    doc_rev_index_list = []
    for rev in item.revision:
        doc_rev_index_list.append(rev_order.index(rev))
    if cur_rev_index < max(doc_rev_index_list):
        print('This REV is SUPERSEEDED')

def set_current(self, item):
    print('set current function start')
    print(item.document)
    doc = item.document
    print('item set file', item.file)
    
    try:
        print(doc.id)
        #print(doc.document_id)
        
        #comme = db.session.query(models.Comments).filter(models.Comments.document_id == doc.id).all()
        #print(comme)
        db.session.query(models.Comments).filter(models.Comments.document_id == doc.id).delete()
        #db.session.query(models.Comments).filter(models.Comments.document_id == item.document_id).delete()
        
        print('not here')
        comments(item)
        print('still here')
        print(doc.revision)
        '''
        for rev in doc.revision:
            print(rev)
            rev.current = False
        item.current = True
        db.session.commit()
        #d.session.close()
        '''
    except:
        print('something wrong in set current')
        pass

def precheck_doc(self, item):
    filename = get_file_original_name(item.file)
    filename_list = filename.split('-',6)

    try:
        filename_list = filename.split('-')
        print(len(filename_list), filename)
        if len(filename_list) != 7:
            return False, "Please check your File Name: " + filename
         
        rev = models.Revisions
        this_reply = False
        rep_extension_list = ['CS REP.xlsx', 'CS REP.xls', 'CS REP.XLS', 'CS REP.XLSX']
        if filename_list[6] in rep_extension_list:
            this_reply = True
        
        print('rev.reply == item.reply',rev.reply, item.reply)
        revision = db.session.query(rev).filter(rev.document_id == item.document_id, rev.revision == item.revision, rev.reply == this_reply).first()
        print(filename_list[6])

        this_rev = item.revision
        if filename_list[6] == 'CS REP.xlsx':
            this_rev = this_rev + '-Reply'
        

        print('revision',revision, 'item rev', this_rev)
        if str(this_rev) == str(revision):
            return False, "The Revision {0} already exist.".format(this_rev)

    except:
        return False, "Something Wrong Whit Your File Name: " + filename

    '''
    try:
        if this_rev != "A":
            last_rev(self, revision)
    except:
        print('something worng with last rev function')
        pass
    '''
    return True, "File uploaded " + filename

def check_Doc(self, item):
    try:
        print('here in check doc', item.file)
        filename = get_file_original_name(item.file)
        filename_list = filename.split('-',6)
        
        
        unit = filename[0:3]
        partner = item.trasmittal[4:7]
        mat = filename[4]
        doctype = filename[6:9]
        serial = filename[10:15]
        sheet = filename[16:19]

        session = db.session
        doc = models.Document
        document = session.query(doc).filter(doc.unit == unit, doc.materialclass == mat,
                                    doc.doctype == doctype, doc.serial == serial, doc.sheet == sheet).first()

        if document:
            #session.close()
            return document.id, document.partner
        
        else:
            document = doc(unit=unit, materialclass=mat, doctype=doctype, 
                        serial=serial, sheet=sheet, partner=partner)
            
            session.add(document)

            session.flush()
            session.commit()
            #session.close()

            return document.id, document.partner
    except:
        abort(400,'Please check your file name. - check_Doc')

def precheck_reply(self,item):
    filename = get_file_original_name(item.file)
    filename_list = filename.split('-',6)
    print('len:',len(filename_list), filename_list[6])
    try:
        if filename_list[6] != 'CS REP.xlsx':
            return False, "Something Wrong Whit Your File Name: " + filename_list[6]
    except:
        return False, "Something is wrong with your file name. " + filename

def check_reply(self, item):
    filename = get_file_original_name(item.file)
    try:
        print('this is the Check reply Functions....')
        print('Reply identification: ', filename[-8:-5] )
        if filename[-8:-5] == "REP":
            return True
    except:
        return False

def check_doc_closed(doc_id):
    doc = models.Document
    document = db.session.query(doc).filter(doc.id==doc_id).first()
    print('Check Doc Closed Function')
    
    closed = True
    for com in document.comments:
        
        if com.closed == False:
            closed = False
            break
        print('comment closed:',com, com.closed)
    document.closed = closed
    document.changed_by_fk = '1'
    db.session.commit()
    #db.session.close()

def set_comments_blank():
    comment = models.Comments
    revision = models.Revisions
    document = models.Document
    
    session = db.session

    revision_reply = session.query(revision).filter(revision.reply == True).all()
    print('the number of revisions in reply is:', len(revision_reply))
    blank_reply = 0
    for rev in revision_reply:
        #session = db.session
        comments_reply = rev.comments

    
        print('the number of comments in reply is:', len(comments_reply))

        for comment in comments_reply:
            
            if comment.reply == ' ' and comment.comment == ' ':
                print('id\t', comment.id, 'autho\t', comment.author, '\tcomm', comment.comment[:15])
                comment.note = 'No comment AND No replies'
                comment.changed_by_fk = '1'
                session.delete(comment)
                blank_reply += 1
                print('*************   NO Comment AND NO Reply   *************')
      
            session.commit()
    
    comment = models.Comments
    comment_all = session.query(comment).all()
    blank = 0
    for comment in comment_all:

        if comment.comment == ' ':
            blank += 1
            print('blank comment')
            session.delete(comment)
    session.commit()
    print('vuoti', blank)
    print('vuoti in reply', blank_reply)
    session.close()
    
def set_comments_included():
    comment = models.Comments
    revision = models.Revisions
    document = models.Document
    
    session = db.session

    # Take only revisions in reply
    revision_reply = session.query(revision).filter(revision.reply == True).all()
    print('the number of revisions in reply is:', len(revision_reply))
    
    # for every revision take the comments
    closed_by_included = 0
    for rev in revision_reply:

        comments_reply = rev.comments
    
        print('the number of comments in reply is:', len(comments_reply))

        # for every comments do things

        for comment in comments_reply:
            print('id\t', comment.id, 'autho\t', comment.author, '\tcomm', comment.comment[:15])
            
            # At this Conditions
            
            if comment.included:
                comment.closed = True
                closed_by_included += 1
                comment.changed_by_fk = '1'
                
                print('*************   Comment Included --> Closed   *************')
            
            
            session.commit()
    print('closed_by_included', closed_by_included)
    session.close()

def timing(f):
    def wrap(*args):
        time1 = time.time()
        ret = f(*args)
        time2 = time.time()
        print('{:s} function took {:.3f} ms'.format(f.__name__, (time2-time1)*1000.0))

        return ret
    return wrap

@timing
def report_all():
    #print('start', time.time())
    file_template = UPLOAD_FOLDER + 'report/cs_dashboard.xlsx'
    filepath = UPLOAD_FOLDER + 'report/' + str(uuid.uuid4()) + 'cs_dashboard.xlsx'
    workbook = openpyxl.load_workbook(file_template)
    
    #comment = models.Comments
    revision = models.Revisions
    document = models.Document

    #wc = workbook['Comments']    
    wr = workbook['Revisions']
    wd = workbook['Documents']

    # 
    # Styles
    #

    ft = Font(bold=True, color='4b1f68')

    # 
    # Populate Comments Sheet
    #

    '''
    comments = db.session.query(comment).all()
    
    row = 1
    col = 1 

    for comment in comments:
        print('comments *********************')
        _ = wc.cell(column=1, row=row, value=str(comment.id))
        _ = wc.cell(column=2, row=row, value=str(comment.partner))
        _ = wc.cell(column=3, row=row, value=str(comment.document))
        _ = wc.cell(column=4, row=row, value=str(comment.revision))
        _ = wc.cell(column=5, row=row, value=str(comment.included))
        _ = wc.cell(column=6, row=row, value=str(comment.closed))

        row += 1
    '''

    # 
    # Populate revisions Sheet
    #

    
    revisions = db.session.query(revision).all()
    start = time.time()
    #print('revision query done', time.time())
    
    _ = wr.cell(column=1, row=1, value='ID')
    _.font = ft
    _ = wr.cell(column=2, row=1, value='FileName')
    _.font = ft
    _ = wr.cell(column=3, row=1, value='Revision')
    _.font = ft
    _ = wr.cell(column=4, row=1, value='Transmittal')
    _.font = ft
    _ = wr.cell(column=5, row=1, value='Trans. Date')
    _.font = ft
    _ = wr.cell(column=6, row=1, value='Response Code')
    _.font = ft
    _ = wr.cell(column=7, row=1, value='Note')
    _.font = ft
    _ = wr.cell(column=8, row=1, value='Current')
    _.font = ft


    row = 2
    col = 1 

    for revision in revisions:
        _ = wr.cell(column=1, row=row, value=str(revision.id))
        _ = wr.cell(column=2, row=row, value=str(revision.file_name()))
        _ = wr.cell(column=3, row=row, value=str(revision.revision))
        _ = wr.cell(column=4, row=row, value=str(revision.trasmittal))
        _ = wr.cell(column=5, row=row, value=str(revision.date_trs))
        _ = wr.cell(column=6, row=row, value=str(revision.action_code))
        _ = wr.cell(column=7, row=row, value=str(revision.note))
        _ = wr.cell(column=8, row=row, value=str(revision.current))

        row += 1
    rev_time = time.time()
    #print('write revision sheet', time.time)
    wr.column_dimensions['A'].width = 5
    wr.column_dimensions['B'].width = 35
    wr.column_dimensions['D'].width = 20
    wr.column_dimensions['F'].width = 35
    wr.column_dimensions['G'].width = 35
    # 
    # Populate documents Sheet
    #

    # 
    # Set the labels for the document sheet
    #
    
    documents = db.session.query(document).all()
    
    
    _ = wd.cell(column=1, row=1, value='ID')
    _.font = ft
    _ = wd.cell(column=2, row=1, value='Bapco Code')
    _.font = ft
    _ = wd.cell(column=3, row=1, value='Partner')
    _.font = ft
    _ = wd.cell(column=4, row=1, value='Revisions')
    _.font = ft
    _ = wd.cell(column=5, row=1, value='Open')
    _.font = ft
    _ = wd.cell(column=6, row=1, value='Included')
    _.font = ft
    _ = wd.cell(column=7, row=1, value='Closed')
    _.font = ft
    
    

    # 
    # Set the rows for the document sheet
    #
    
    row = 2
    

    for document in documents:
        doc = [document.id, document.name(), document.partner, str(document.revision), document.count_open(), document.count_included(), document.count_closed()]
        wd.append(doc)
        
        '''
        _ = wd.cell(column=1, row=row, value=str(document.id))
        _ = wd.cell(column=2, row=row, value=str(document.name()))
        _ = wd.cell(column=3, row=row, value=str(document.partner))
        _ = wd.cell(column=4, row=row, value=str(document.revision))
        _ = wd.cell(column=5, row=row, value=str(document.count_open()))
        _ = wd.cell(column=6, row=row, value=str(document.count_included()))
        _ = wd.cell(column=7, row=row, value=str(document.count_closed()))
     
        '''

        row += 1
    
    doc_time = time.time()
    #print('write document sheet', time.time())
    print('write revision took {:.3f} ms'.format((rev_time-start)*1000.0))
    print('write document took {:.3f} ms'.format((doc_time-rev_time)*1000.0))

    wd.column_dimensions['A'].width = 5
    wd.column_dimensions['B'].width = 20
    wd.column_dimensions['D'].width = 35
    wd.column_dimensions['E'].width = 8
    wd.column_dimensions['F'].width = 8
    wd.column_dimensions['G'].width = 8

    wd.page_setup.fitToWidth = 1
    
    tab = Table(displayName='Documents', ref='A1:E'+ str(row))
    style = TableStyleInfo(name='TableStyleMedium9')
    tab.tableStyleInfo = style
    #wd.add_table(tab)
    
    workbook.save(filepath) 
    
    file = filepath
    #print(file.name)
    db.session.close()
    
    return file



def check_doc_closed2():
    doc = models.Document
    documents = db.session.query(doc).all()
    
    for document in documents:

        print('Check Doc Function')
        closed = True
        for com in document.comments:
            
            if com.closed == False:
                closed = False
                break
            print('comment closed:',com, com.closed)
        document.closed = closed
        document.changed_by_fk = '1'
        db.session.commit()
    db.session.close()
        #return True
                                      
def report_url(self):
    url = 'https://report.quasarpm.com/superset/explore/?form_data=%7B%22datasource%22%3A%2262__table%22%2C%22viz_type%22%3A%22pivot_table%22%2C%22slice_id%22%3A152%2C%22granularity_sqla%22%3A%22TransmDate%22%2C%22time_grain_sqla%22%3A%22P1M%22%2C%22since%22%3A%221+years+ago%22%2C%22until%22%3A%22now%22%2C%22groupby%22%3A%5B%22Year%22%2C%22BapcoResponse%22%5D%2C%22columns%22%3A%5B%22IsReply%22%5D%2C%22metrics%22%3A%5B%22Total+CS%22%5D%2C%22row_limit%22%3A10000%2C%22pandas_aggfunc%22%3A%22sum%22%2C%22pivot_margins%22%3Atrue%2C%22number_format%22%3A%22.3s%22%2C%22combine_metric%22%3Afalse%2C%22where%22%3A%22%22%2C%22having%22%3A%22%22%2C%22filters%22%3A%5B%7B%22val%22%3A%5B1%5D%2C%22op%22%3A%22in%22%2C%22col%22%3A%22IsReply%22%7D%5D%2C%22url_params%22%3A%7B%7D%7D&standalone=true&height=400'
    out = urlparse(url)
    params = parse_qs(out.query)
    form_data = json.loads(params['form_data'][0])
    url_only = out._replace(query=None).geturl()
    print('url only', url_only)   

    url_new = requests.get(url_only, params=params)
    print('urln new', url_new.url, url_new.content )  
    print('FORM DATA ARGS:')
    for k,v in form_data.items():
        
        print(k, '--> ',v)
       
    
    return print(params['form_data'])